import openpyxl
from collections import defaultdict


def work_book():
    wb = openpyxl.load_workbook(filename='Классный журнал 7а.xlsx')
    return wb


def empty_line(a):
    wb = work_book()
    sheet = wb[a]
    for i in range(1, 1048576):
        if sheet.cell(row=i, column=2).value == None:
            break
    return i


def student_list(a):
    wb = work_book()
    sheet = wb[a]
    i = empty_line(a)
    for j in range(1, i):
        print(str(sheet.cell(row=j, column=1).value) +
              " " + str(sheet.cell(row=j, column=2).value))


def subject_list(a):
    wb = work_book()
    sheet = wb[a]
    i = empty_line(a)
    for j in range(1, i):
        print(str(sheet.cell(row=j, column=1).value) +
              " " + str(sheet.cell(row=j, column=2).value))


def mass(a):
    wb = work_book()
    sheet = wb[a]
    i = empty_line(a)
    return [{sheet.cell(row=j, column=1).value: (sheet.cell(row=j, column=2).value)} for j in range(2, i-1)]


def printer(a):
    wb = work_book()
    sheet = wb[a]
    i = empty_line(a)
    return [[sheet.cell(row=j, column=1).value, sheet.cell(row=j, column=2).value, sheet.cell(row=j, column=3).value, sheet.cell(row=j, column=4).value]for j in range(2, i)]

    # arr=defaultdict(list)
    # for j in range(2,i):
    #     arr[sheet.cell(row=j, column=1).value].append([sheet.cell(row=j, column=2).value,sheet.cell(row=j, column=3).value,sheet.cell(row=j, column=4).value])
    # return dict(arr)

# # print(printer("Посещаемость"))
# print(printer("Оценки"))


